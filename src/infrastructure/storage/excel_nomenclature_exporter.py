import shutil
import tempfile
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

from src.config.settings import Settings, get_settings
from src.domain.entities import Product
from src.domain.exceptions import ExportError

from .column_map_registry import COLUMN_MAPPING

IMAGES_DIRNAME = "images"
NOMENCLATURE_FILENAME = "nomenclature.xlsx"


class ExcelNomenclatureExporter:
    """Экспорт карточек товаров в Excel по шаблону номенклатуры."""

    def __init__(self, settings: Settings | None = None) -> None:
        self._settings = settings or get_settings()

    def export(self, products: list[Product], output_path: Path) -> Path:
        """Экспортирует карточки товаров в ZIP-архив с номенклатурой.

        Parameters
        ----------
        products : list[Product]
            Карточки товаров для экспорта.
        output_path : Path
            Путь к выходному ZIP-архиву.

        Returns
        -------
        Path
            Абсолютный путь к сохранённому ZIP-архиву.

        Raises
        ------
        ValueError
            Если список ``products`` пуст.
        ExportError
            Если Excel-шаблон номенклатуры не найден.
        """
        if not products:
            raise ValueError("products list must not be empty")

        template_path = self._settings.resolve_path(
            self._settings.paths.nomenclature_template
        )
        if not template_path.exists():
            raise ExportError(
                str(output_path),
                f"Template not found: {template_path}"
            )

        resolved_output = self._settings.resolve_path(output_path)
        resolved_output.parent.mkdir(parents=True, exist_ok=True)

        with tempfile.TemporaryDirectory() as tmp_dir:
            workdir = Path(tmp_dir)

            images_dir = workdir / IMAGES_DIRNAME
            images_dir.mkdir(parents=True, exist_ok=True)

            workbook = load_workbook(template_path)
            sheet = workbook.active
            row = sheet.max_row + 1

            for product in products:
                image_rel_path = self._copy_image(product, images_dir)
                self._write_row(
                    sheet=sheet,
                    row=row,
                    row_data=self._product_to_row(product, image_rel_path)
                )
                row += 1

            workbook.save(workdir / NOMENCLATURE_FILENAME)
            self._build_archive(workdir, resolved_output)

            return resolved_output.resolve()

    @staticmethod
    def _copy_image(product: Product, images_dir: Path) -> str | None:
        """Копирует изображение товара в директорию ``images/`` архива.

        Parameters
        ----------
        product : Product
            Карточка товара.
        images_dir : Path
            Директория ``images/`` во временной рабочей директории.

        Returns
        -------
        str | None
            Относительный путь к изображению внутри архива, либо ``None``,
            если у товара нет изображения или файл недоступен.
        """
        if product.image_path is None or not product.image_path.exists():
            return None

        destination = images_dir / product.image_path.name
        shutil.copy2(product.image_path, destination)

        return f"{IMAGES_DIRNAME}/{destination.name}"

    @staticmethod
    def _product_to_row(
        product: Product,
        image_rel_path: str | None,
    ) -> dict[str, object | None]:
        specs = product.specifications

        def _val(name: str) -> float | None:
            measurement = getattr(specs, name, None)
            return measurement.value if measurement is not None else None

        def _unit(name: str) -> str | None:
            measurement = getattr(specs, name, None)
            return measurement.unit if measurement is not None else None

        def _flag(name: str) -> bool:
            return _val(name) is not None

        return {
            "name": product.name,
            "sku": product.sku,
            "weight_unit": _unit("weight"),
            "weight_flag": _flag("weight"),
            "weight_value": _val("weight"),
            "length_unit": _unit("length"),
            "length_flag": _flag("length"),
            "length_value": _val("length"),
            "description": product.description_text,
            "brand": product.brand,
            "manufacturer": product.manufacturer,
            "image_path": image_rel_path,
            "volume_unit": _unit("volume"),
            "volume_flag": _flag("volume"),
            "volume_value": _val("volume"),
            "square_unit": _unit("square"),
            "square_flag": _flag("square"),
            "square_value": _val("square"),
            "origin_country": None,
        }

    @staticmethod
    def _write_row(sheet, row: int, row_data: dict[str, object | None]) -> None:
        for col_key, value in row_data.items():
            col = COLUMN_MAPPING.get(col_key)
            if col is not None:
                sheet[f"{col}{row}"] = value

    @staticmethod
    def _build_archive(workdir: Path, output_path: Path) -> None:
        """Упаковывает содержимое рабочей директории в ZIP-архив.

        Parameters
        ----------
        workdir : Path
            Рабочая директория с файлом номенклатуры и папкой ``images/``.
        output_path : Path
            Путь к выходному ZIP-архиву.
        """
        with ZipFile(output_path, mode="w", compression=ZIP_DEFLATED) as archive:
            for file_path in workdir.rglob("*"):
                if file_path.is_file():
                    archive.write(
                        file_path,
                        arcname=file_path.relative_to(workdir),
                    )
